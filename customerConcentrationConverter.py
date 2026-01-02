#!/usr/bin/env python3
"""
Customer Concentration (Sales by Customer Summary) Converter
Converts CSV, XLSX, and PDF Sales by Customer Summary reports to simplified JSON array format
"""

import json
import csv
import sys
import os
from pathlib import Path
import argparse
from typing import Dict, List, Any, Optional

# Try to import optional dependencies
try:
    import openpyxl
    XLSX_SUPPORT = True
except ImportError:
    XLSX_SUPPORT = False

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False


class CustomerConcentrationConverter:
    """Converts Sales by Customer Summary reports to simplified JSON array format"""
    
    def parse_amount(self, value: str) -> float:
        """Parse monetary amount from string"""
        if not value or value.strip() == '':
            return 0.0
        # Remove currency symbols, commas, and whitespace
        clean_value = value.replace('$', '').replace(',', '').replace('"', '').strip()
        try:
            return float(clean_value)
        except ValueError:
            return 0.0
    
    def calculate_percentages(self, customers: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Calculate percentage of total for each customer"""
        # Calculate grand total
        grand_total = sum(c['revenue'] for c in customers)
        
        # Calculate percentages
        for customer in customers:
            if grand_total > 0:
                customer['percentage'] = (customer['revenue'] / grand_total) * 100
            else:
                customer['percentage'] = 0.0
        
        # Sort by revenue descending
        customers.sort(key=lambda x: x['revenue'], reverse=True)
        
        return customers
    
    def parse_csv(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse CSV file and convert to simplified JSON array format"""
        customers = []
        customer_map = {}  # Track parent-child relationships
        
        with open(filepath, 'r', encoding='utf-8') as f:
            # Skip header lines
            for _ in range(4):
                f.readline()
            
            reader = csv.DictReader(f)
            current_parent = None
            
            for row in reader:
                customer_name = row.get('Customer', '').strip()
                
                # Skip empty rows or TOTAL row
                if not customer_name or customer_name.upper() == 'TOTAL':
                    break
                
                # Check if this is a "Total for" row (sub-customer total)
                if customer_name.startswith('Total for '):
                    # Extract parent name
                    parent_name = customer_name.replace('Total for ', '')
                    if parent_name in customer_map:
                        # Update the parent's total from the "Total for" row
                        total = self.parse_amount(row.get('Total', '0'))
                        customer_map[parent_name]['revenue'] = total
                    current_parent = None
                    continue
                
                # Parse total amount
                total = self.parse_amount(row.get('Total', '0'))
                
                # Check if this is a parent customer (has no amount or will have sub-items)
                # In the CSV, parent customers appear with no amount, followed by their sub-customers
                if total == 0.0:
                    # This might be a parent customer
                    current_parent = customer_name
                    customer_map[customer_name] = {
                        'customerName': customer_name,
                        'revenue': 0.0,
                        'percentage': 0.0
                    }
                    continue
                
                # If we're in a parent context, this is a sub-customer - add to parent's total
                if current_parent and current_parent in customer_map:
                    customer_map[current_parent]['revenue'] += total
                else:
                    # Regular customer
                    if customer_name not in customer_map:
                        customer_map[customer_name] = {
                            'customerName': customer_name,
                            'revenue': total,
                            'percentage': 0.0
                        }
        
        # Convert map to list
        customers = list(customer_map.values())
        
        return self.calculate_percentages(customers)
    
    def parse_xlsx(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse XLSX file and convert to simplified JSON array format"""
        if not XLSX_SUPPORT:
            raise ImportError("openpyxl is required for XLSX support. Install with: pip install openpyxl")
        
        customers = []
        customer_map = {}
        
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        # Find header row
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and any('Customer' in str(cell) for cell in row if cell):
                header_row = idx
                break
        
        if not header_row:
            raise ValueError("Could not find header row in XLSX file")
        
        # Get column indices
        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header}
        
        current_parent = None
        
        # Parse data rows
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[col_map.get('Customer', 0)]:
                continue
            
            customer_name = str(row[col_map.get('Customer', 0)]).strip()
            
            if customer_name.upper() == 'TOTAL':
                break
            
            if customer_name.startswith('Total for '):
                parent_name = customer_name.replace('Total for ', '')
                if parent_name in customer_map:
                    total = self.parse_amount(str(row[col_map.get('Total', 1)] or '0'))
                    customer_map[parent_name]['revenue'] = total
                current_parent = None
                continue
            
            total = self.parse_amount(str(row[col_map.get('Total', 1)] or '0'))
            
            if total == 0.0:
                current_parent = customer_name
                customer_map[customer_name] = {
                    'customerName': customer_name,
                    'revenue': 0.0,
                    'percentage': 0.0
                }
                continue
            
            if current_parent and current_parent in customer_map:
                customer_map[current_parent]['revenue'] += total
            else:
                if customer_name not in customer_map:
                    customer_map[customer_name] = {
                        'customerName': customer_name,
                        'revenue': total,
                        'percentage': 0.0
                    }
        
        customers = list(customer_map.values())
        
        return self.calculate_percentages(customers)
    
    def parse_pdf(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse PDF file and convert to simplified JSON array format"""
        if not PDF_SUPPORT:
            raise ImportError("pdfplumber is required for PDF support. Install with: pip install pdfplumber")
        
        customers = []
        customer_map = {}
        
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                # Try table extraction first
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        # Find header row
                        header_row_idx = -1
                        for i, row in enumerate(table):
                            if row and any(cell and 'Customer' in str(cell) for cell in row):
                                header_row_idx = i
                                break
                        
                        if header_row_idx == -1:
                            continue
                        
                        current_parent = None
                        
                        # Process data rows
                        for row in table[header_row_idx + 1:]:
                            if not row or not row[0]:
                                continue
                            
                            customer_name = str(row[0]).strip()
                            
                            if customer_name.upper() == 'TOTAL':
                                break
                            
                            if customer_name.startswith('Total for '):
                                parent_name = customer_name.replace('Total for ', '')
                                if parent_name in customer_map:
                                    total = self.parse_amount(str(row[1] if len(row) > 1 else '0'))
                                    customer_map[parent_name]['revenue'] = total
                                current_parent = None
                                continue
                            
                            total = self.parse_amount(str(row[1] if len(row) > 1 else '0'))
                            
                            if total == 0.0:
                                current_parent = customer_name
                                customer_map[customer_name] = {
                                    'customerName': customer_name,
                                    'revenue': 0.0,
                                    'percentage': 0.0
                                }
                                continue
                            
                            if current_parent and current_parent in customer_map:
                                customer_map[current_parent]['revenue'] += total
                            else:
                                if customer_name not in customer_map:
                                    customer_map[customer_name] = {
                                        'customerName': customer_name,
                                        'revenue': total,
                                        'percentage': 0.0
                                    }
        
        customers = list(customer_map.values())
        
        return self.calculate_percentages(customers)
    
    def convert_file(self, filepath: Path) -> List[Dict[str, Any]]:
        """Convert a file to JSON array format based on its extension"""
        ext = filepath.suffix.lower()
        
        if ext == '.csv':
            return self.parse_csv(filepath)
        elif ext == '.xlsx':
            return self.parse_xlsx(filepath)
        elif ext == '.pdf':
            return self.parse_pdf(filepath)
        else:
            raise ValueError(f"Unsupported file format: {ext}")
    
    def convert_to_json(self, filepath: Path, output_path: Optional[Path] = None) -> str:
        """Convert a file to JSON format"""
        result = self.convert_file(filepath)
        
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2)
            return f"Converted {len(result)} customers to {output_path}"
        else:
            return json.dumps(result, indent=2)


def main():
    parser = argparse.ArgumentParser(description='Convert Sales by Customer Summary reports to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')
    parser.add_argument('--batch', action='store_true', help='Process all files in a directory')
    
    args = parser.parse_args()
    
    converter = CustomerConcentrationConverter()
    
    if args.batch:
        input_path = Path(args.input)
        if not input_path.is_dir():
            print(f"Error: {input_path} is not a directory", file=sys.stderr)
            sys.exit(1)
        
        output_dir = Path(args.output) if args.output else input_path.parent / 'converted'
        output_dir.mkdir(exist_ok=True)
        
        for file in input_path.glob('*'):
            if file.suffix.lower() in ['.csv', '.xlsx', '.pdf']:
                try:
                    output_file = output_dir / f"{file.stem}_customer_concentration.json"
                    result = converter.convert_to_json(file, output_file)
                    print(result)
                except Exception as e:
                    print(f"Error processing {file}: {e}", file=sys.stderr)
    else:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: {input_path} does not exist", file=sys.stderr)
            sys.exit(1)
        
        try:
            if args.output:
                result = converter.convert_to_json(input_path, Path(args.output))
                print(result)
            else:
                print(converter.convert_to_json(input_path))
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)


if __name__ == '__main__':
    main()
