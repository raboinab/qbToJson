#!/usr/bin/env python3
"""
Vendor Concentration (Expenses by Vendor Summary) Converter
Converts CSV, XLSX, and PDF Expenses by Vendor Summary reports to simplified JSON array format
"""

import json
import csv
import sys
import os
from pathlib import Path
import argparse
from typing import Dict, List, Any, Optional

# Try to import optional dependencies
try:
    import openpyxl
    XLSX_SUPPORT = True
except ImportError:
    XLSX_SUPPORT = False

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False


class VendorConcentrationConverter:
    """Converts Expenses by Vendor Summary reports to simplified JSON array format"""
    
    def parse_amount(self, value: str) -> float:
        """Parse monetary amount from string"""
        if not value or value.strip() == '':
            return 0.0
        # Remove currency symbols, commas, and whitespace
        clean_value = value.replace('$', '').replace(',', '').replace('"', '').strip()
        try:
            return float(clean_value)
        except ValueError:
            return 0.0
    
    def calculate_percentages(self, vendors: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Calculate percentage of total for each vendor"""
        # Calculate grand total
        grand_total = sum(v['payments'] for v in vendors)
        
        # Calculate percentages
        for vendor in vendors:
            if grand_total > 0:
                vendor['percentage'] = (vendor['payments'] / grand_total) * 100
            else:
                vendor['percentage'] = 0.0
        
        # Sort by payments descending
        vendors.sort(key=lambda x: x['payments'], reverse=True)
        
        return vendors
    
    def parse_csv(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse CSV file and convert to simplified JSON array format"""
        vendors = []
        
        with open(filepath, 'r', encoding='utf-8') as f:
            # Skip header lines
            for _ in range(4):
                f.readline()
            
            reader = csv.DictReader(f)
            for row in reader:
                vendor_name = row.get('Vendor', '').strip()
                
                # Skip TOTAL row
                if vendor_name.upper() == 'TOTAL':
                    break
                
                # Skip empty vendor names
                if not vendor_name:
                    continue
                
                # Parse total amount
                total = self.parse_amount(row.get('Total', '0'))
                
                vendors.append({
                    'vendorName': vendor_name,
                    'payments': total,
                    'percentage': 0.0  # Will be calculated later
                })
        
        return self.calculate_percentages(vendors)
    
    def parse_xlsx(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse XLSX file and convert to simplified JSON array format"""
        if not XLSX_SUPPORT:
            raise ImportError("openpyxl is required for XLSX support. Install with: pip install openpyxl")
        
        vendors = []
        
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        # Find header row
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and any('Vendor' in str(cell) for cell in row if cell):
                header_row = idx
                break
        
        if not header_row:
            raise ValueError("Could not find header row in XLSX file")
        
        # Get column indices
        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header}
        
        # Parse data rows
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[col_map.get('Vendor', 0)]:
                continue
            
            vendor_name = str(row[col_map.get('Vendor', 0)]).strip()
            
            if vendor_name.upper() == 'TOTAL':
                break
            
            total = self.parse_amount(str(row[col_map.get('Total', 1)] or '0'))
            
            vendors.append({
                'vendorName': vendor_name,
                'payments': total,
                'percentage': 0.0
            })
        
        return self.calculate_percentages(vendors)
    
    def parse_pdf(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse PDF file and convert to simplified JSON array format"""
        if not PDF_SUPPORT:
            raise ImportError("pdfplumber is required for PDF support. Install with: pip install pdfplumber")
        
        vendors = []
        
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                # Try table extraction first
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        # Find header row
                        header_row_idx = -1
                        for i, row in enumerate(table):
                            if row and any(cell and 'Vendor' in str(cell) for cell in row):
                                header_row_idx = i
                                break
                        
                        if header_row_idx == -1:
                            continue
                        
                        # Process data rows
                        for row in table[header_row_idx + 1:]:
                            if not row or not row[0]:
                                continue
                            
                            vendor_name = str(row[0]).strip()
                            
                            if vendor_name.upper() == 'TOTAL':
                                break
                            
                            total = self.parse_amount(str(row[1] if len(row) > 1 else '0'))
                            
                            vendors.append({
                                'vendorName': vendor_name,
                                'payments': total,
                                'percentage': 0.0
                            })
        
        return self.calculate_percentages(vendors)
    
    def convert_file(self, filepath: Path) -> List[Dict[str, Any]]:
        """Convert a file to JSON array format based on its extension"""
        ext = filepath.suffix.lower()
        
        if ext == '.csv':
            return self.parse_csv(filepath)
        elif ext == '.xlsx':
            return self.parse_xlsx(filepath)
        elif ext == '.pdf':
            return self.parse_pdf(filepath)
        else:
            raise ValueError(f"Unsupported file format: {ext}")
    
    def convert_to_json(self, filepath: Path, output_path: Optional[Path] = None) -> str:
        """Convert a file to JSON format"""
        result = self.convert_file(filepath)
        
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2)
            return f"Converted {len(result)} vendors to {output_path}"
        else:
            return json.dumps(result, indent=2)


def main():
    parser = argparse.ArgumentParser(description='Convert Expenses by Vendor Summary reports to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')
    parser.add_argument('--batch', action='store_true', help='Process all files in a directory')
    
    args = parser.parse_args()
    
    converter = VendorConcentrationConverter()
    
    if args.batch:
        input_path = Path(args.input)
        if not input_path.is_dir():
            print(f"Error: {input_path} is not a directory", file=sys.stderr)
            sys.exit(1)
        
        output_dir = Path(args.output) if args.output else input_path.parent / 'converted'
        output_dir.mkdir(exist_ok=True)
        
        for file in input_path.glob('*'):
            if file.suffix.lower() in ['.csv', '.xlsx', '.pdf']:
                try:
                    output_file = output_dir / f"{file.stem}_vendor_concentration.json"
                    result = converter.convert_to_json(file, output_file)
                    print(result)
                except Exception as e:
                    print(f"Error processing {file}: {e}", file=sys.stderr)
    else:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: {input_path} does not exist", file=sys.stderr)
            sys.exit(1)
        
        try:
            if args.output:
                result = converter.convert_to_json(input_path, Path(args.output))
                print(result)
            else:
                print(converter.convert_to_json(input_path))
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)


if __name__ == '__main__':
    main()
