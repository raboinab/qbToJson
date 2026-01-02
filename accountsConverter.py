#!/usr/bin/env python3
"""
Chart of Accounts Document Converter
Converts CSV, XLSX, and PDF account lists to QuickBooks JSON format
Specifically designed for Chart of Accounts reports
"""

import json
import csv
import sys
import os
from datetime import datetime, timezone
from pathlib import Path
import argparse
import re
from typing import Dict, List, Any, Optional

# Try to import optional dependencies
try:
    import openpyxl
    XLSX_SUPPORT = True
except ImportError:
    XLSX_SUPPORT = False

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False


class AccountsConverter:
    """Converts Chart of Accounts documents to QuickBooks-style JSON format"""
    
    
    def __init__(self):
        self.next_id = 200  # Starting ID for generated accounts
        
    def generate_account_id(self) -> str:
        """Generate a unique account ID"""
        id_str = str(self.next_id)
        self.next_id += 1
        return id_str
    
    def get_classification_from_type(self, type_str: str) -> str:
        """Determine classification based on type string"""
        type_lower = type_str.lower()
        if 'equity' in type_lower:
            return 'EQUITY'
        elif 'expense' in type_lower:
            return 'EXPENSE'
        elif 'income' in type_lower:
            return 'REVENUE'
        elif 'asset' in type_lower:
            return 'ASSET'
        elif 'liabilit' in type_lower:
            return 'LIABILITY'
        else:
            return 'EXPENSE'  # Default fallback
    
    def get_account_type_from_type(self, type_str: str) -> str:
        """Determine account type based on type string"""
        type_lower = type_str.lower()
        if 'equity' in type_lower:
            return 'EQUITY'
        elif 'other expense' in type_lower:
            return 'OTHER_EXPENSE'
        elif 'expense' in type_lower:
            return 'EXPENSE'
        elif 'other income' in type_lower:
            return 'OTHER_INCOME'
        elif 'income' in type_lower:
            return 'INCOME'
        elif 'other current asset' in type_lower:
            return 'OTHER_CURRENT_ASSET'
        elif 'fixed asset' in type_lower:
            return 'FIXED_ASSET'
        elif 'asset' in type_lower:
            return 'OTHER_CURRENT_ASSET'
        elif 'current liabilit' in type_lower:
            return 'CURRENT_LIABILITY'
        elif 'long term liabilit' in type_lower:
            return 'LONG_TERM_LIABILITY'
        elif 'liabilit' in type_lower:
            return 'OTHER_CURRENT_LIABILITY'
        else:
            return 'EXPENSE'  # Default fallback
    
    def create_account_object(self, name: str, type_str: str, detail_type: str, 
                            description: Optional[str] = None, 
                            balance: float = 0.0) -> Dict[str, Any]:
        """Create a QuickBooks-style account object"""
        
        # Get classification and account type based on the type string
        classification = self.get_classification_from_type(type_str)
        account_type = self.get_account_type_from_type(type_str)
        
        # Use detail type as-is (just clean it up a bit)
        # Remove special characters and normalize spacing
        account_subtype = detail_type.strip()
        # Remove forward slashes and replace with empty string
        account_subtype = account_subtype.replace('/', '')
        # Replace ampersands with 'And'
        account_subtype = account_subtype.replace('&', 'And')
        # Replace spaces with no spaces to match QuickBooks format
        account_subtype = account_subtype.replace(' ', '')
        
        # Generate timestamp
        timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000+00:00')
        
        return {
            "id": self.generate_account_id(),
            "syncToken": "0",
            "metaData": {
                "createdByRef": None,
                "createTime": timestamp,
                "lastModifiedByRef": None,
                "lastUpdatedTime": timestamp,
                "lastChangedInQB": None,
                "synchronized": None
            },
            "customField": [],
            "attachableRef": [],
            "domain": "QBO",
            "status": None,
            "sparse": False,
            "name": name,
            "subAccount": False,
            "parentRef": None,
            "description": description,
            "fullyQualifiedName": name,
            "accountAlias": None,
            "txnLocationType": None,
            "active": True,
            "classification": classification,
            "accountType": account_type,
            "accountSubType": account_subtype,
            "accountPurposes": [],
            "acctNum": None,
            "acctNumExtn": None,
            "bankNum": None,
            "openingBalance": None,
            "openingBalanceDate": None,
            "currentBalance": balance,
            "currentBalanceWithSubAccounts": balance,
            "currencyRef": {
                "value": "USD",
                "name": "United States Dollar",
                "type": None
            },
            "taxAccount": None,
            "taxCodeRef": None,
            "onlineBankingEnabled": None,
            "journalCodeRef": None,
            "accountEx": None,
            "finame": None
        }
    
    def parse_csv(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse CSV file and convert to account objects"""
        accounts = []
        
        with open(filepath, 'r', encoding='utf-8') as f:
            # Skip header lines
            for _ in range(3):  # Skip first 3 lines based on sample
                f.readline()
            
            reader = csv.DictReader(f)
            for row in reader:
                # Skip empty rows, total row, or metadata rows
                full_name = row.get('Full name', '')
                if not full_name or full_name == 'TOTAL':
                    continue
                
                # Skip rows that look like report metadata (contain dates/times)
                if any(keyword in full_name.lower() for keyword in ['basis', 'gmtz', 'accrual', 'cash']):
                    continue
                
                # Parse balance
                balance_str = row.get('Total balance', '0').replace('$', '').replace(',', '')
                try:
                    balance = float(balance_str) if balance_str else 0.0
                except ValueError:
                    balance = 0.0
                
                account = self.create_account_object(
                    name=row['Full name'],
                    type_str=row.get('Type', ''),
                    detail_type=row.get('Detail type', ''),
                    description=row.get('Description'),
                    balance=balance
                )
                accounts.append(account)
        
        return accounts
    
    def parse_xlsx(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse XLSX file and convert to account objects"""
        if not XLSX_SUPPORT:
            raise ImportError("openpyxl is required for XLSX support. Install with: pip install openpyxl")
        
        accounts = []
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        # Find header row
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and 'Full name' in str(row):
                header_row = idx
                break
        
        if not header_row:
            raise ValueError("Could not find header row in XLSX file")
        
        # Get column indices
        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_map = {header: idx for idx, header in enumerate(headers) if header}
        
        # Parse data rows
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[col_map.get('Full name', 0)]:
                continue
            
            name = str(row[col_map.get('Full name', 0)])
            if name == 'TOTAL':
                continue
            
            # Skip rows that look like report metadata (contain dates/times)
            if any(keyword in name.lower() for keyword in ['basis', 'gmtz', 'accrual', 'cash']):
                continue
            
            # Parse balance
            balance_str = str(row[col_map.get('Total balance', 4)] or '0')
            balance_str = balance_str.replace('$', '').replace(',', '')
            try:
                balance = float(balance_str) if balance_str else 0.0
            except ValueError:
                balance = 0.0
            
            account = self.create_account_object(
                name=name,
                type_str=row[col_map.get('Type', 1)] or '',
                detail_type=row[col_map.get('Detail type', 2)] or '',
                description=row[col_map.get('Description', 3)],
                balance=balance
            )
            accounts.append(account)
        
        return accounts
    
    def parse_pdf(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse PDF file and convert to account objects"""
        if not PDF_SUPPORT:
            raise ImportError("pdfplumber is required for PDF support. Install with: pip install pdfplumber")
        
        accounts = []
        
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                
                lines = text.split('\n')
                
                # Find header line
                header_idx = -1
                for i, line in enumerate(lines):
                    # Look for headers in uppercase
                    if 'FULL NAME' in line and 'TYPE' in line and 'DETAIL TYPE' in line:
                        header_idx = i
                        break
                
                if header_idx == -1:
                    continue
                
                # Process data lines after header
                for i in range(header_idx + 1, len(lines)):
                    line = lines[i].strip()
                    
                    if not line or line.startswith('TOTAL'):
                        break
                    
                    # Skip metadata lines
                    if any(keyword in line.lower() for keyword in ['basis', 'gmtz', 'accrual', 'cash']):
                        continue
                    
                    # Parse the line - this is tricky because columns aren't clearly delimited
                    # We'll use a heuristic approach based on known patterns
                    
                    # Known account types that help us split the line
                    type_patterns = ['Equity', 'Expenses', 'Income', 'Other Current Assets', 
                                   'Other Expense', 'Other Income']
                    
                    # Find where the type starts
                    type_start = -1
                    found_type = None
                    for pattern in type_patterns:
                        idx = line.find(pattern)
                        if idx > 0:  # Must not be at the beginning
                            type_start = idx
                            found_type = pattern
                            break
                    
                    if type_start > 0:
                        # Extract name (everything before type)
                        name = line[:type_start].strip()
                        
                        # Extract the rest after type
                        rest = line[type_start + len(found_type):].strip()
                        
                        # Try to extract detail type - it's usually the next part
                        # For simplicity, we'll take everything up to a potential balance
                        # or use the whole rest if no balance
                        detail_type = rest
                        
                        # Remove any trailing balance info (usually ends with numbers or $)
                        import re
                        detail_type = re.sub(r'\s*\$?[\d,]+\.?\d*\s*$', '', detail_type).strip()
                        
                        if not detail_type:
                            # Use a default based on type
                            if 'Expense' in found_type:
                                detail_type = 'Other Miscellaneous Service Cost'
                            elif 'Income' in found_type:
                                detail_type = 'Service/Fee Income'
                            elif 'Asset' in found_type:
                                detail_type = 'Other Current Assets'
                            else:
                                detail_type = 'Other'
                        
                        account = self.create_account_object(
                            name=name,
                            type_str=found_type,
                            detail_type=detail_type,
                            balance=0.0
                        )
                        accounts.append(account)
        
        return accounts
    
    def convert_file(self, filepath: Path) -> List[Dict[str, Any]]:
        """Convert a file to account objects based on its extension"""
        ext = filepath.suffix.lower()
        
        if ext == '.csv':
            return self.parse_csv(filepath)
        elif ext == '.xlsx':
            return self.parse_xlsx(filepath)
        elif ext == '.pdf':
            return self.parse_pdf(filepath)
        else:
            raise ValueError(f"Unsupported file format: {ext}")
    
    def convert_to_json(self, filepath: Path, output_path: Optional[Path] = None) -> str:
        """Convert a file to JSON format"""
        accounts = self.convert_file(filepath)
        
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(accounts, f, indent=2)
            return f"Converted {len(accounts)} accounts to {output_path}"
        else:
            return json.dumps(accounts, indent=2)


def main():
    parser = argparse.ArgumentParser(description='Convert account documents to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')
    parser.add_argument('--batch', action='store_true', help='Process all files in a directory')
    
    args = parser.parse_args()
    
    converter = AccountsConverter()
    
    if args.batch:
        # Process all files in directory
        input_path = Path(args.input)
        if not input_path.is_dir():
            print(f"Error: {input_path} is not a directory", file=sys.stderr)
            sys.exit(1)
        
        output_dir = Path(args.output) if args.output else input_path.parent / 'converted'
        output_dir.mkdir(exist_ok=True)
        
        for file in input_path.glob('*'):
            if file.suffix.lower() in ['.csv', '.xlsx', '.pdf']:
                try:
                    output_file = output_dir / f"{file.stem}.json"
                    result = converter.convert_to_json(file, output_file)
                    print(result)
                except Exception as e:
                    print(f"Error processing {file}: {e}", file=sys.stderr)
    else:
        # Process single file
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: {input_path} does not exist", file=sys.stderr)
            sys.exit(1)
        
        try:
            if args.output:
                result = converter.convert_to_json(input_path, Path(args.output))
                print(result)
            else:
                print(converter.convert_to_json(input_path))
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)


if __name__ == '__main__':
    main()
